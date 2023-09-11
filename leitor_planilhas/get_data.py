import openpyxl


path = 'C:/Users/Usuario/Documents/Erich/GitHub/proforco/leitor_planilhas'
wb = openpyxl.load_workbook(path+'/PAINEL DE AULAS.xlsx', data_only=True)
sheets = wb.sheetnames
sheet = wb['FIXO SEGUNDA']
for i in range(1,sheet.max_column+1):
    celula = str(sheet.cell(1, i).value)
    if celula.upper().startswith('ERICH'):
        break


cols = [i, i+2]
linhas = []
for j in range(2, sheet.max_row+1):
    linha = [sheet.cell(j, cols[0]).value,
            sheet.cell(j, cols[1]).value]
    
    if str(linha[0]).strip() == 'NAO TEVE AULA': break
    elif linha[0] != None and linha[1] != None:
        linha[1] = linha[1].isoformat()
        linhas.append(linha)


for p in linhas:
    print(p)
        