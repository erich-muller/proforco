# COLETA OS DADOS DE UMA DETERMINADA PLANILHA
# E RETORNA UMA LISTA DE DOIS ELEMENTOS SENDO O
# PRIMEIRO O NOME DO ALUNO E O SEGUNDO O HORÁRIO
 
import openpyxl


class Planilha():
    def __init__(self, path:str) -> None:
        self.path = path
        try:
            self.wb = openpyxl.load_workbook(path, data_only=True)
            self.sheetnames = self.wb.sheetnames
            
            sheets = {}
            for sheetname in self.sheetnames:
                sheets[sheetname] = self.wb[sheetname]
            self.sheets = sheets  
        except:
            print('Não foi possível obter a planilha')
            raise ImportError
    
    def get_alunos(self, sheetname:str, professor:str='ERICH'):
        try:
            sheet = self.wb[sheetname]
        except:
            print(f'Página {sheetname} não lançada, tentando fixo.')
            try: 
                sheet = self.wb['FIXO '+sheetname]
            except:
                print('Não foi possível encontrar', 'FIXO '+sheetname)
                raise ImportError

        # ENCONTRAR O INTERVALO DE COLS DO PROFESSOR
        for i in range(1,sheet.max_column+1):
            celula = str(sheet.cell(1, i).value)
            if celula.upper().startswith(professor.upper()):
                break
        cols = [i, i+2]
        
        # ENCONTRAR O INTERVALO DE LINHAS DO PROFESSOR
        linhas = []
        for j in range(2, sheet.max_row+1):
            linha = [sheet.cell(j, cols[0]).value,
                    sheet.cell(j, cols[1]).value]
            if str(linha[0]).strip() == 'NAO TEVE AULA': break
            elif linha[0] != None and linha[1] != None:
                
                # VERIFICAR SE SÃO DUAS AULAS CONSECUTIVAS
                if sheet.cell(j+1, cols[0]).value != linha[0]:
                    linha[0], linha[1] = linha[0].strip(), linha[1].isoformat()
                    linhas.append(linha)
        
        return linhas

